#!/usr/bin/env python3
"""
Inspect Excel file to extract indentation from background colors
"""

import openpyxl
from openpyxl.styles import Color
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')

# Check all Excel files
excel_files = [
    'EPD_DataSet.xlsx',
    'ILCD_Format_Documentation_v1.3_reformatted_final_2025-09-12.xlsx',
    'TEWOG N234_Digital Data Requirements V2.0 Content_2025-09-12_ed.xlsx'
]

for excel_file in excel_files:
    excel_path = os.path.join(DATA_DIR, excel_file)
    if not os.path.exists(excel_path):
        print(f"File not found: {excel_file}")
        continue
    
    print(f"\n{'='*80}")
    print(f"Inspecting: {excel_file}")
    print(f"{'='*80}\n")
    
    wb = openpyxl.load_workbook(excel_path)
    
    # Find the correct sheet
    sheet_name = None
    for name in wb.sheetnames:
        if 'ILCD' in name or 'EPD' in name or 'Format' in name or 'Doc' in name:
            sheet_name = name
            break
    
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    
    ws = wb[sheet_name]
    print(f"Sheet name: {sheet_name}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Find the header row by checking multiple columns
    header_row = None
    for i in range(1, min(15, ws.max_row + 1)):
        # Check multiple potential header columns
        for col in [1, 2, 3, 4, 5]:
            cell_value = ws.cell(i, col).value
            if cell_value and ('Element' in str(cell_value) or 'Field' in str(cell_value) or 'Name' in str(cell_value)):
                header_row = i
                break
        if header_row:
            break
    
    if not header_row:
        print("Showing first 10 rows to identify header:")
        for i in range(1, min(11, ws.max_row + 1)):
            row_preview = " | ".join([str(ws.cell(i, col).value)[:30] if ws.cell(i, col).value else "" for col in range(1, 6)])
            print(f"  Row {i}: {row_preview}")
        header_row = 1  # Default to first row
    
    print(f"\nUsing row {header_row} as header")
    
    # Print headers
    print("\nHeaders:")
    for col in range(1, min(15, ws.max_column + 1)):
        header = ws.cell(header_row, col).value
        if header:
            print(f"  Col {col}: {header}")
    
    # Find the "Element/Attribute Name" column
    element_col = None
    for col in range(1, min(20, ws.max_column + 1)):
        header_value = ws.cell(header_row, col).value
        if header_value and 'Element' in str(header_value) and 'Attribute' in str(header_value):
            element_col = col
            break
    
    if not element_col:
        element_col = 3  # Default to column 3
    
    print(f"\nElement/Attribute Name column: {element_col}")
    
    # Analyze first 30 data rows for color patterns
    print(f"\nFirst 30 data rows (colors and values):")
    print(f"{'Row':<5} {'Color':<20} {'Element/Attribute Name':<50}")
    print("-" * 80)
    
    color_to_indent = {}
    
    for i in range(header_row + 1, min(header_row + 31, ws.max_row + 1)):
        # Get the Element/Attribute Name cell's background color
        element_name_cell = ws.cell(i, element_col)
        fill = element_name_cell.fill
        
        color_rgb = None
        if fill and fill.start_color:
            if isinstance(fill.start_color.rgb, str):
                color_rgb = fill.start_color.rgb
            elif hasattr(fill.start_color, 'index') and fill.start_color.index:
                color_rgb = f"INDEX:{fill.start_color.index}"
        
        # Get element name
        element_name = element_name_cell.value if element_name_cell.value else ""
        
        print(f"{i:<5} {str(color_rgb):<20} {str(element_name):<50}")
        
        # Track unique colors
        if color_rgb and color_rgb not in color_to_indent:
            color_to_indent[color_rgb] = len(color_to_indent)
    
    print(f"\n\nUnique colors found: {len(color_to_indent)}")
    for color, idx in sorted(color_to_indent.items(), key=lambda x: x[1]):
        print(f"  {idx}: {color}")
    
    wb.close()

print("\n" + "="*80)
print("Inspection complete!")
