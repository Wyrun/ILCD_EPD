#!/usr/bin/env python3
"""
Check the 'order' column to see if it contains hierarchical information
"""

import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
EXCEL_FILE = 'ILCD_Format_Documentation_v1.3_reformatted_final_2025-09-12.xlsx'

wb = openpyxl.load_workbook(os.path.join(DATA_DIR, EXCEL_FILE))
ws = wb.active

print(f"{'Row':<5} {'Order':<12} {'Color':<12} {'Element/Attribute Name':<50}")
print("-" * 80)

for i in range(2, 40):
    order_val = ws.cell(i, 1).value  # Column 1 = order
    element_val = ws.cell(i, 9).value  # Column 9 = Element/Attribute Name
    
    color_rgb = "None"
    cell = ws.cell(i, 9)
    if cell.fill and cell.fill.start_color:
        if isinstance(cell.fill.start_color.rgb, str):
            color_rgb = cell.fill.start_color.rgb[:8]  # Show first 8 chars
    
    if element_val:
        print(f"{i:<5} {str(order_val):<12} {color_rgb:<12} {str(element_val):<50}")

wb.close()

print("\n\nObservations:")
print("- Order column appears to use hierarchical numbering (1, 1.1, 1.1.1, etc.)")
print("- Count the dots in 'order' to determine indent level!")
print("  Example: '1' = indent 0, '1.1' = indent 1, '1.1.1' = indent 2")
