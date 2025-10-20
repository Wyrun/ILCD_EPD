#!/usr/bin/env python3
"""
Analyze the color hierarchy in Excel file to build indent mapping
"""

import openpyxl
import os
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')

# Use the file that the conversion script references
EXCEL_FILE = 'ILCD_Format_Documentation_v1.3_reformatted_final_2025-09-12.xlsx'
excel_path = os.path.join(DATA_DIR, EXCEL_FILE)

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print(f"Analyzing: {EXCEL_FILE}")
print(f"Sheet: {ws.title}\n")

# Find header row and element column
header_row = 1
element_col = 9  # From previous inspection

# Collect all colors and their associated element names
color_examples = defaultdict(list)

for i in range(header_row + 1, min(header_row + 100, ws.max_row + 1)):
    cell = ws.cell(i, element_col)
    element_name = cell.value if cell.value else ""
    
    color_rgb = None
    if cell.fill and cell.fill.start_color:
        if isinstance(cell.fill.start_color.rgb, str):
            color_rgb = cell.fill.start_color.rgb
    
    if color_rgb and element_name:
        color_examples[color_rgb].append((i, element_name))

print("Colors found with examples:\n")
print(f"{'Color':<15} {'Count':<8} {'Examples'}")
print("-" * 80)

for color, examples in sorted(color_examples.items(), key=lambda x: len(x[1]), reverse=True):
    ex_str = ", ".join([f"{name[:30]}" for row, name in examples[:5]])
    print(f"{color:<15} {len(examples):<8} {ex_str}")

# Manually analyze structure by looking at known hierarchy
print("\n\nAnalyzing hierarchical structure:")
print("-" * 80)

# Look for patterns like processDataSet > processInformation > dataSetInformation > UUID
test_rows = [
    (2, "processDataSet"),
    (7, "processInformation"),
    (8, "dataSetInformation"),
    (9, "UUID"),
    (12, "functionalUnitFlowProperties"),
    (15, "classificationInformation"),
    (16, "classification"),
    (17, "@name"),
    (19, "class"),
]

print(f"{'Row':<5} {'Color':<15} {'Element Name':<40}")
print("-" * 65)

for row_num, expected_name in test_rows:
    cell = ws.cell(row_num, element_col)
    actual_name = cell.value if cell.value else ""
    
    color_rgb = "None"
    if cell.fill and cell.fill.start_color:
        if isinstance(cell.fill.start_color.rgb, str):
            color_rgb = cell.fill.start_color.rgb
    
    print(f"{row_num:<5} {color_rgb:<15} {actual_name:<40}")

print("\n\nSuggested color-to-indent mapping:")
print("-" * 80)

# Based on the visual hierarchy described by the user
color_to_indent = {
    'FFFFC000': 0,  # Orange - root level (processDataSet, processInformation)
    'FFFFD783': 1,  # Lighter orange - second level (dataSetInformation)
    'FFFFF3D9': 2,  # Pale orange - third level (UUID, name, synonyms)
    'FFFFFFCC': 2,  # Yellow - also third level (some alternate styling)
    'FFD7F7BD': 2,  # Green - third level EPD24 namespace
    '00000000': 3,  # White/no fill - fourth level (attributes like @name)
}

for color, indent in sorted(color_to_indent.items(), key=lambda x: x[1]):
    count = len(color_examples.get(color, []))
    print(f"  '{color}': {indent}  # {count} occurrences")

wb.close()

print("\n" + "="*80)
print("Analysis complete!")
print("\nNext step: Update convert_xlsx_to_adoc.py to use this color mapping")
